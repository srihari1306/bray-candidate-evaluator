"""
Export service for CSV and Excel file generation.
"""

import io
from typing import Optional

from app.models.schemas import CandidateResult
from app.utils.logger import get_logger

logger = get_logger("export")


class ExportService:
    """Service for exporting candidate data to CSV and Excel formats."""

    def export_to_csv(self, candidates: list[CandidateResult], skill_names: list[str] = None) -> bytes:
        """Export candidate results to CSV format."""
        import csv

        output = io.StringIO()
        writer = csv.writer(output)

        # Header
        headers = ["Rank", "Candidate", "Email", "Overall Score", "JD Match", "Recommendation"]
        if skill_names:
            headers.extend(skill_names)
        headers.extend(["Missing Skills", "Summary"])
        writer.writerow(headers)

        # Data rows
        for i, c in enumerate(candidates, 1):
            row = [i, c.candidate_name, c.email, c.overall_score, c.jd_match_score, c.overall_recommendation.value]
            if skill_names:
                skill_map = {s.skill: s.score for s in c.skill_scores}
                for sn in skill_names:
                    row.append(skill_map.get(sn, "N/A"))
            row.extend(["; ".join(c.missing_skills), c.summary])
            writer.writerow(row)

        return output.getvalue().encode("utf-8")

    def export_to_excel(self, candidates: list[CandidateResult], skill_names: list[str] = None) -> bytes:
        """Export candidate results to Excel format."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Candidate Rankings"

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        good_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        bad_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        # Headers
        headers = ["Rank", "Candidate", "Email", "Overall Score", "JD Match", "Recommendation"]
        if skill_names:
            headers.extend(skill_names)
        headers.extend(["Missing Skills", "Strengths", "Summary"])

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Data
        for i, c in enumerate(candidates, 1):
            row = i + 1
            ws.cell(row=row, column=1, value=i).border = thin_border
            ws.cell(row=row, column=2, value=c.candidate_name).border = thin_border
            ws.cell(row=row, column=3, value=c.email).border = thin_border

            # Overall score with conditional formatting
            score_cell = ws.cell(row=row, column=4, value=c.overall_score)
            score_cell.border = thin_border
            score_cell.alignment = Alignment(horizontal="center")
            if c.overall_score >= 75:
                score_cell.fill = good_fill
            elif c.overall_score >= 50:
                score_cell.fill = warn_fill
            else:
                score_cell.fill = bad_fill

            ws.cell(row=row, column=5, value=c.jd_match_score).border = thin_border
            ws.cell(row=row, column=6, value=c.overall_recommendation.value).border = thin_border

            col_offset = 7
            if skill_names:
                skill_map = {s.skill: s.score for s in c.skill_scores}
                for j, sn in enumerate(skill_names):
                    score = skill_map.get(sn, 0)
                    cell = ws.cell(row=row, column=col_offset + j, value=score)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center")
                    if score >= 75:
                        cell.fill = good_fill
                    elif score >= 50:
                        cell.fill = warn_fill
                    else:
                        cell.fill = bad_fill
                col_offset += len(skill_names)

            ws.cell(row=row, column=col_offset, value="; ".join(c.missing_skills)).border = thin_border
            ws.cell(row=row, column=col_offset + 1, value="; ".join(c.strengths)).border = thin_border
            ws.cell(row=row, column=col_offset + 2, value=c.summary).border = thin_border

        # Auto-width columns
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()


def get_export_service() -> ExportService:
    return ExportService()
